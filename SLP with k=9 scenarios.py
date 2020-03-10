import time
start_time = time.time()
from pulp import *
from openpyxl import load_workbook

wb=load_workbook(filename= 'Compiled output.xlsx')
ws=wb['kmeans']
ws1=wb['price data']

M=10000000
Cap=1000000
cin=305000
cout=-457500
d=1
Horizon=24
N=9

T=list(range(Horizon))
S=list(range(N))

p=[[0 for t in range(Horizon)] for s in range(N)]
q=[0  for s in range(N)]

for t in range(Horizon):
    for s in range(N):
        p[s][t]=ws.cell(row=t+6, column=s+38).value
        
for s in range(N):
    q[s]=ws.cell(row=31, column=s+38).value
         
prob= LpProblem("EL-NR", LpMaximize)
x= LpVariable.matrix("x", (S,T), None, None, LpContinuous)
I= LpVariable.matrix("I", (S,T), 0, None, LpContinuous)

prob+= lpSum([lpSum([-d*x[s][t]*p[s][t]*q[s] for t in T]) for s in S])

for s in S:
    for t in T:
        prob+=cout<=x[s][t]
        prob+=x[s][t]<=cin
        prob+=I[s][t]<=Cap
        if t>0:
            prob+=I[s][t]==I[s][t-1]+x[s][t-1]
    prob+= I[s][Horizon-1]+x[s][Horizon-1]==I[s][0]
    prob+= I[s][0]==0  
prob.solve()     
  
print(prob.objective.value())
for v in prob.variables():
    if v.name[:1]=='x':
        print(v.name, "=", v.varValue)  
        
print("--- %s seconds ---" % (time.time() - start_time))
#Xlist=list()
#for s in list(range(N)):
#    Ylist=list()
#    for t in list(range(Horizon)):
#        for v in prob.variables():
#            if v.name[0:]=='x_'+str(s)+'_'+str(t):
#                Ylist.append(v.name)
#    Xlist.append(Ylist)
#           
#print(Xlist) 
#print(len(Xlist)) 


#for v in prob.variables():
#    print(v.name, "=", v.varValue)        
print("--- %s seconds ---" % (time.time() - start_time))
   

