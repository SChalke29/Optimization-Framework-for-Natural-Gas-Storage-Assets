import time
start_time = time.time()
from pulp import *
from openpyxl import load_workbook

wb=load_workbook(filename= 'Compiled output.xlsx')
ws=wb['kmeans']         # For 9 scenarios

M=10000000
Cap=1000000
cin=305000
cout=-457500
d=1
Horizon=24
N=962

T=list(range(Horizon))
S=list(range(N))

p=[[0 for t in range(Horizon)] for s in range(N)]
q=[0  for s in range(N)]

for t in range(Horizon):
    for s in range(N):
        p[s][t]=ws.cell(row=t+6, column=s+38).value    # For 9 scenarios
                 
prob= LpProblem("EL-NR", LpMaximize)

x= LpVariable.matrix("x", (T), None, None, LpContinuous)
I= LpVariable.matrix("I", (T), 0, None, LpContinuous)
v= LpVariable('v',  None, None, LpContinuous)

prob+= v

for s in S:
    prob+= v<=lpSum([-d*x[t]*p[s][t] for t in T]) 

for t in T:
    prob+=cout<=x[t]
    prob+=x[t]<=cin
    prob+=I[t]<=Cap
    if t>0:
        prob+=I[t]==I[t-1]+x[t-1]

prob+= I[Horizon-1]+x[Horizon-1]==I[0]
prob+= I[0]==0  
prob.solve()

print(prob.objective.value())

for v in prob.variables():
    if v.name[:1]=="x":
        print(v.name, "=", v.varValue)
            
print("--- %s seconds ---" % (time.time() - start_time))
   

