import time
start_time = time.time()
from pulp import *
from openpyxl import load_workbook
import numpy as np
import pandas as pd

wb=load_workbook(filename= 'Price Curves 12-5-18.xlsx')
ws1=wb['Sheet2']

M=10000000
Cap=1000000
cin=305000
cout=-457500
d=1
Horizon=24
N=962

T=list(range(Horizon))
S=list(range(N))

p=[[0 for t in range(Horizon)] for s in range(N)]
q=[0 for s in range(N)]

for t in range(Horizon):
    for s in range(N):
        p[s][t]=ws1.cell(row=t+2, column=s+2).value
        
for s in range(N):
    q[s]=(1/N)
         
prob= LpProblem("EL-NR", LpMaximize)
x= LpVariable.matrix("x", (S,T), None, None, LpContinuous)
I= LpVariable.matrix("I", (S,T), 0, None, LpContinuous)

prob+= lpSum([lpSum([-d*x[s][t]*p[s][t]*q[s] for t in T]) for s in S])
for s in S:
    for t in T:
        prob+=cout<=x[s][t]
        prob+=x[s][t]<=cin
        prob+=I[s][t]<=Cap
        if t>0:
            prob+=I[s][t]==I[s][t-1]+x[s][t-1]
    prob+= I[s][Horizon-1]+x[s][Horizon-1]==I[s][0]
    prob+= I[s][0]==0  
prob.solve(GUROBI())

#ylist=list() 
#for s in S:
#    for v in prob.variables():
#        if v.name[:3]=='x_'+str(s):
#            print(v.name, "=", v.varValue)
#            ylist.append(v.varValue)
    
#print([ylist[x:x+24] for x in range(0, len(ylist), 24)])
print(prob.objective.value())

#df= pd.DataFrame(ylist[x:x+24] for x in range(0, len(ylist), 24)).transpose()
#writer = pd.ExcelWriter('Xst of SLP.xlsx', engine='xlsxwriter') 
#df.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=0, header=False, index=False)
#writer.save() 

print("--- %s seconds ---" % (time.time() - start_time))

   

