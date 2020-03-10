import time
start_time = time.time()
from pulp import *
from openpyxl import load_workbook
import pandas as pd

#For 962 clusters#
wb1=load_workbook(filename= 'Price Curves 12-5-18.xlsx')
ws1=wb1['Sheet2']

M=10000000
Cap=1000000
cin=305000
cout=-457500
d=1
Horizon=24
N=962 #For 962 Clusters

T=list(range(Horizon))
S=list(range(N))

p=[0 for t in range(Horizon)] 
OFlist=list()
ylist=list()
for s in S:
    for t in range(Horizon):
        p[t]=ws1.cell(row=t+2, column=2+s).value     #For 962 Clusters
    prob= LpProblem("EL-NR", LpMaximize)
    x= LpVariable.matrix("x", (T), None, None, LpContinuous)
    I= LpVariable.matrix("I", (T), 0, None, LpContinuous)
              
    prob+= (lpSum([-p[t]*x[t] for t in T]))
    for t in T:
        prob+=cout<=x[t]
        prob+=x[t]<=cin
        prob+=I[t]<=Cap
        if t>0:
            prob+=I[t]==I[t-1]+x[t-1]
    
        prob+= I[Horizon-1]+x[Horizon-1]==I[0]
        prob+= I[0]==0  
        prob.solve(GUROBI())
         
    xlist=list()
    for v in prob.variables():
        if v.name[:1]=="x":
            y= v.name, "=", v.varValue
            xlist.append(v.varValue)
    ylist.append(xlist)
    OFlist.append(prob.objective.value())

#print(OFlist)
print(min(OFlist))

df1=pd.DataFrame(OFlist).transpose()
df= pd.DataFrame(ylist).transpose()
writer = pd.ExcelWriter('X_Sol values for 962 clusters.xlsx', engine='xlsxwriter') #For 962 Clusters
df.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=0, header=False, index=False)
df1.to_excel(writer, sheet_name='Sheet2', startcol=0, startrow=0, header=False, index=False)
writer.save() 
            
print("--- %s seconds ---" % (time.time() - start_time))

