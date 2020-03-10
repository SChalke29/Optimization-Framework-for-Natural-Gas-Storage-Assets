import time
start_time = time.time()
from pulp import *
from openpyxl import load_workbook
import pandas as pd

wb=load_workbook(filename= 'Compiled output.xlsx')
ws=wb['kmeans']

M=100000000000000000
Cap=1000000
cin=305000
cout=-457500
d=1

Horizon=24

T=list(range(24))
S=list(range(9))

#p=[[0 for j in range(Horizon)] for i in range(N)]
p=[[0 for t in range(24)] for s in range(9)]
q=[0  for s in range(9)]

for t in range(24):
    for s in range(9):
        p[s][t]=ws.cell(row=t+6, column=s+38).value

for s in range(9):
    q[s]=ws.cell(row=31, column=s+38).value
         

#alphalist=[0, 0.05, 0.1, 0.15, 0.20, 0.25, 0.35, 0.40, 0.45, 0.50, 0.55, 0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90, 0.95, 1]

OFlist2=list()
b=520000
while b <= 554000: 
    OFlist1=list()
    alpha=0
    while alpha <=1:
        b=round(b,2)
        alpha=round(alpha,2)
        prob= LpProblem("EL-NR", LpMaximize)
        zz = LpVariable.matrix("zz", (S), 0,1, LpInteger)
        x= LpVariable.matrix("x", (S,T), None, None, LpContinuous)
        I= LpVariable.matrix("I", (S,T), 0, None, LpContinuous)
    
        prob+= lpSum([lpSum([-d*x[s][t]*p[s][t]*q[s] for t in T]) for s in S])
        
        for s in S:
            for t in T:
                prob+=cout<=x[s][t]
                prob+=x[s][t]<=cin
                prob+=I[s][t]<=Cap
                if t>0:
                    prob+=I[s][t]==I[s][t-1]+x[s][t-1]
            prob+= I[s][Horizon-1]+x[s][Horizon-1]==I[s][0]
            prob+= I[s][0]==0  
        for s in S:
            prob+=lpSum([-d*x[s][t]*p[s][t] for t in T])-b>= -M*(1-zz[s])
        prob+=lpSum([q[s]*zz[s] for s in S])>=alpha
        prob.solve(GUROBI()) 
        OFlist1.append(prob.objective.value())
        
        alpha+=0.05
    b+=100
    OFlist2.append(OFlist1)
    
#print(OFlist2)
df=pd.DataFrame(OFlist2).transpose()
writer = pd.ExcelWriter('Chance1.xlsx', engine='xlsxwriter')
df.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=0, header=False, index=True)
writer.save()

print("--- %s seconds ---" % (time.time() - start_time))
#print(alphalist)

#for v in prob.variables():
#    if v.varValue>0:
#        if v.name[:0]=="zz":
#            print(v.name, "=", v.varValue)
#    print(v.name, "=", v.varValue)
#            print(v.name)
#        if v.name[:1]=="w":
#            print(v.name, "=", v.varValue)
            
print("--- %s seconds ---" % (time.time() - start_time))
   

