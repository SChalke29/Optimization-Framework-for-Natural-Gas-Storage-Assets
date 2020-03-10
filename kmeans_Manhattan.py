#matrix math
import numpy as np
#graphing
import matplotlib.pyplot as plt
#graphing animation
import matplotlib.animation as animation

from scipy.spatial.distance import squareform, pdist
import scipy
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict
import math

#load textfile dataset (2D data points)
#For each user, how many packets are sent per second and what's the size of packet 
#anamolies (DDOS attempts) will have lots of big packets sent in a short amount of time
def load_dataset(name):
    return np.loadtxt(name)

#euclidian distance between 2 data points. For as many data points as necessary.
def euclidian(a, b):
#    return np.linalg.norm(a-b)  #euclidean
    return np.linalg.norm(a-b,1)  #manhattan

#lets define a plotting algorithm for our dataset and our centroids
#def plot(dataset, history_centroids, belongs_to):
#    #we'll have 2 colors for each centroid cluster
#    colors = ['r','g','b','y','p','o','black','brown','pink','gray']
#    #split our graph by its axis and actual plot
#    fig, ax = plt.subplots()
#    
#    #for each point in our dataset 
#    for index in range(dataset.shape[0]):
#        #get all the points assigned to a cluster
#        instances_close = [i for i in range(len(belongs_to)) if belongs_to[i] == index]
#        #assign each data point in that cluster a color and plot it 
#        for instance_index in instances_close:
#            ax.plot(dataset[instance_index][0], dataset[instance_index][1], (colors[index] + 'o'))
#            
#    #lets also log the history of centroids calculated via training 
#    history_points = []
#    #for each centroid ever calculated 
#    for index, centroids in enumerate(history_centroids):
#        #print them all out 
#        for inner, item in enumerate(centroids):
#            if index == 0:
#                history_points.append(ax.plot(item[0], item[1], 'bo')[0])
#            else:
#                history_points[inner].set_data(item[0], item[1])
##                print("centroids {} {}".format(index, item))
#                
##                plt.show()
#                plt.pause(0.8)

def kmeans(k, epsilon=0, distance='euclidian'):
    #list to store past centroid
    history_centroids = []
    #set the distance calculation type
    if distance == 'euclidian':
        dist_method = euclidian
    #set the dataset
    dataset = load_dataset('durudataset.txt')
    # dataset = dataset[:, 0:dataset.shape[1] - 1]
    #get the number of rows (instances) and columns (features) from the dataset
    num_instances, num_features = dataset.shape
    #define k centroids (how many clusters do we want to find?) chosen randomly
    prototypes = dataset[np.random.randint(0, num_instances - 1, size=k)]
    #set these to our list of past centroid (to show progress over time)
    history_centroids.append(prototypes)
    #to keep track of centroid at every iteration
    prototypes_old = np.zeros(prototypes.shape)
    #to store clusters
    belongs_to = np.zeros((num_instances, 1))
    norm = dist_method(prototypes, prototypes_old)
    iteration = 0
    while norm > epsilon:
        iteration += 1
        norm = dist_method(prototypes, prototypes_old)
        prototypes_old = prototypes
        #for each instance in the dataset
        for index_instance, instance in enumerate(dataset):
            #define a distance vector of size k
            dist_vec = np.zeros((k, 1))
            #for each centroid 
            for index_prototype, prototype in enumerate(prototypes):
                #compute the distance between x(data points) and centroid 
                dist_vec[index_prototype] = dist_method(prototype, instance)
            #find the smallest distance, assign that distance to a cluster     
            belongs_to[index_instance, 0] = np.argmin(dist_vec)

        tmp_prototypes = np.zeros((k, num_features))
        
        #for each cluster, k of them 
        for index in range(len(prototypes)):
            #get all the points assigned to a cluster
            instances_close = [i for i in range(len(belongs_to)) if belongs_to[i] == index]
            #find the mean of those points, this is our new centroid 
            prototype = np.mean(dataset[instances_close], axis=0)
            # prototype = dataset[np.random.randint(0, num_instances, size=1)[0]]
            #add out new centroid to our new temporary list
            tmp_prototypes[index, :] = prototype
        
        #set the new list to the current list
        prototypes = tmp_prototypes
        
        #add our calculated centroids to our history of plotting
        history_centroids.append(tmp_prototypes)

    # plot(dataset, history_centroids, belongs_to)
    #return calculated centroids, history of them all, and assignments for 
    return prototypes, history_centroids, belongs_to

#main file
def execute():
    #load dataset
    clusters=20
    for z in list(range(1,clusters+1)):
        
        dataset = load_dataset('durudataset.txt')
        #train the model on the data
        centroids, history_centroids, belongs_to = kmeans(z)
        
        wb=load_workbook(filename= 'pricedata.xlsx')
        ws=wb['Sheet1']
        p=[[0 for j in range(24)] for i in range(962)]
        for i in range(962):
            for j in range(24):
                p[i][j]=ws.cell(row=i+2, column=j+2).value
        
        B= list(belongs_to.flatten())
        C=[]
        for b in B:
            C.append(int(b))
    
        d = defaultdict(list)
        for key, value in zip(C, p):
            d[key].append(value)
        cluster_numbers=list()
        withinness=list()
        cluster_size=list()
        for m in list(range(z)):
            xxxx=list()
            yyyy=list()
            for k in range(len(dict(d)[m])):
                H=[abs(j-i) for i, j in zip(history_centroids[-1][m], dict(d)[m][k])]
                J=sum(H)
                xxxx.append(H)
                yyyy.append(J)
            cluster_size.append(len(dict(d)[m]))
            withinness.append(sum(yyyy))
#        cluster_numbers.append(belongs_to.flatten().transpose())
#        print(withinness)
#        print(sum(withinness))
#        print(history_centroids[-1])
#        print(cluster_size)
##        print(cluster_numbers)
#        print('---------------------------------------------------------------------')
        
        df1= pd.DataFrame(data=withinness)
        df2= pd.DataFrame(data=[sum(withinness)])
        df3= pd.DataFrame(data=history_centroids[-1].flatten())
        df4= pd.DataFrame(data=cluster_size)
        df5= pd.DataFrame(data=belongs_to.flatten())
#        print(df1)
#        print(df2)
#        print(df3)
#        print(df4)
#        print(df5)
#        print('---------------------------------------------------------------------')
        
        writer = pd.ExcelWriter('C:\Sixth Term @Dal\k_means_clustering-master\kmeans_manhattan clusters\\'+str('Cluster')+str(z)+'.xlsx', engine='xlsxwriter')
        df1.to_excel(writer, sheet_name='cluster_withinnesss', startcol=0, startrow=0, header=False, index=False)
        df2.to_excel(writer, sheet_name='Total_withinness', startcol=0, startrow=0, header=False, index=False)
        df3.to_excel(writer, sheet_name='centroids', startcol=0, startrow=0, header=False, index=False)
        df4.to_excel(writer, sheet_name='cluster_size', startcol=0, startrow=0, header=False, index=False)
        df5.to_excel(writer, sheet_name='belongs_to', startcol=0, startrow=0, header=False, index=False)
        writer.save()
           
#        plot the results 
#        plot(dataset, history_centroids, belongs_to)
   
#do everything
execute()