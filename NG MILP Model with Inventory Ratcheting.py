
####################################Natural Gas Asset Optimization Problem Including Inventory Ratcheting#######################################
#Importing the required modules
import time
start_time = time.time()
from pulp import *
from openpyxl import load_workbook
from pulp import solvers
import pandas as pd

#Creating the lists to be exported to MS Excel file through pandas dataframe
vName= list() 
vValue= list() 
wValue= list()
iValue= list()
gValue= list()
Actions=list()

#Time horizon
Horizon=24
#Number of inventory ratchets
pieces=3

#######################Defining sets##################################################
T=list(range(Horizon)) #Time periods
L=list(range(pieces)) #Inventory ratchets

#######################Defining the parameters#########################################
#parameter-1.Natural Gas price data
wb=load_workbook(filename= '1.Price Curves 12-5-18.xlsx')
ws=wb['Sheet1']
p=[0 for j in range(Horizon)] 
for j in range(Horizon):
    p[j]=ws.cell(row=965, column=j+2).value
    
#parameter-2.Inventory Upper bounds
UB=[0 for l in range(pieces)]
UB[0]=150001
UB[1]=300001
UB[2]=1000000

#parameter-3.Inventory Lower bounds
LB=[0 for l in range(pieces)]
LB[0]=0
LB[1]=150000
LB[2]=300000

#parameter-4.Max. monthly injection quantity
f=[0 for l in range(pieces)]
f[0]=10000*30.5
f[1]=8000*30.5
f[2]=6000*30.5

#parameter-5.Max. monthly withdrawal quantity
g=[0 for l in range(pieces)]
g[0]=4000*30.5
g[1]=8000*30.5
g[2]=15000*30.5

Cap=1000000 #parameter-6.Maximum asset capacity
M=Cap #parameter-7.Big M for binary variable constraints
d=.97 #parameter-8.Loss factor

##############################Defining the variables#############################################################################
v = LpVariable.matrix("v", (T),0, None, LpContinuous) #Variable-1 to 24.Injection quantity during time t 
w = LpVariable.matrix("w", (T),0, None, LpContinuous) #Variable-25 to 48.Withdrawal quantity during time t 
I = LpVariable.matrix("I", (T),0, None, LpInteger) #Variable-49 to 72.Inventory in hand at the beginning of time t 
z = LpVariable.matrix("z", (T), 0, 1, LpBinary) #Variable-73 to 96.Inventory in hand at the beginning of time t 
lb= LpVariable.matrix("lambda", (T,L,L), 0, 1, LpContinuous) #Variable-97 to 312.maximum proportion of ratchet k used for injection in time t, when starting in ratchet l
mu= LpVariable.matrix("mu", (T,L,L), 0, 1, LpContinuous) #Variable-313 to 528 .maximum proportion of ratchet k used for withdrawal in time t, when starting in ratchet l 
r=  LpVariable.matrix("r", (T,L), 0 ,1, LpBinary) #Variable-529 to 600.(Binary Variable: 1 if LB[l] ≤ I[t] ≤ UB[l], 0 otherwise)
zz= LpVariable.matrix("zz", (T,L,L),0, 1,LpBinary) #Variable-601 to 816.Binary variables used for selecting proper ratchet order when injecting
xx= LpVariable.matrix("xx", (T,L,L),0,1,LpBinary) #Variable-817 to 1032.Binary variables used for selecting proper ratchet order when withdrawing 

######################################Objective function (Max. expected profit)########################################################
prob= LpProblem("EL-NR", LpMaximize)
prob+= (lpSum([p[t]*(-v[t]+w[t]) for t in T]))

########################################Defining the constraints#########################################################################
for t in T:
    prob+= w[t]<=I[t] #Constraint-1.Limiting the withdrawal quantity to be less than or equal to the inventory in hand
    prob+= w[t]<=M*z[t] #Constraint-2.Injection and withdrawal do not occur during the same time interval
    prob+= v[t]<=M*(1-z[t]) #Constraint-3.Injection and withdrawal do not occur during the same time interval
    prob+= I[t]<=Cap #Constraint-4.Inventory can not exceed the maximum asset capacity
    prob+= lpSum([r[t][l] for l in L])==1 #Constraint-5.Ensures that the time t begins in only one ratchet
              
for t in T:
    prob+= I[t]<=lpSum([r[t][l]*UB[l] for l in L]) #Constraint-6.Inventory in hand lies within the Upper & Lower bound of inventory ratchet
    prob+= I[t]>=lpSum([r[t][l]*LB[l] for l in L]) #Constraint-7.Inventory in hand lies within the Upper & Lower bound of inventory ratchet

def constraint_1(t,L):
    value = 0
    for l in L:
        for k in L:
            if k>=l:
                value+= lb[t][l][k]*f[k]
    return value
    
def constraint_2(t,L):
    value = 0
    for l in L:
        for k in L:
            if k<=l:
                value+= mu[t][l][k]*g[k]
    return value

for t in T:
    prob+=v[t]<= constraint_1(t,L) #Constraint-8.Injection quantity for the month strictly follows the rate of monthly injection for all ratchets
    prob+=w[t]<= constraint_2(t,L) #Constraint-9.withdrawal quantity for the month strictly follows the rate of monthly injection for all ratchets
  
#Constraint-10.Max. proportion of injection ratchet l used when ratchet remains the same (i.e.l=k) during time t                
for t in T:
    for l in L[:pieces-1]:
        for k in L:
            if k==l:
                prob+=lb[t][l][k]<=((UB[l]-I[t])/f[k])+M*(1-r[t][l]) 
                
#Constraint-11.Max. proportion of injection ratchet l used when ratchet changes (i.e.l!=k) during time t                 
for t in T:
    for l in L:
        for k in L:
            if k>l:
                prob+=lb[t][l][k]<=((UB[k]-UB[k-1])/f[k])+M*(1-r[t][l]) 
                
#Constraint-12. total injection ratcheting proportions equals to 1 or 0 if no ratchet change during time t          
for t in T:
    for l in L:
        for k in L:
            if k>=l:
                prob+= lpSum([lb[t][l][k] for k in L])==r[t][l] 

#Constraint-13.Max. proportion of withdrawal ratchet l used when ratchet remains the same (i.e.l=k) during time t 
for t in T:
    for l in L[1:]:
        for k in L:
            if k==l:
                prob+=mu[t][l][k]<=((I[t]-LB[l])/g[l])+M*(1-r[t][l]) 

#Constraint-14.Max. proportion of withdrawal ratchet l used when ratchet changes (i.e.l!=k) during time t 
for t in T:
    for l in L:
        for k in L:
            if k<l:
                prob+=mu[t][l][k]<=((LB[k+1]-LB[k])/g[k])+M*(1-r[t][l])
 
#Constraint-15. total withdrawal ratcheting proportions equals to 1 or 0 if no ratchet change during time t              
for t in T:
    for l in L:
        for k in L:
            if k<=l:
                prob+= lpSum([mu[t][l][k] for k in L])==r[t][l]

#Constraint-16. Selecting proper ratcheting order for injection                 
for t in T:
    for l in L:
        for k in L:
            if k>=l:
                prob+=lb[t][l][k]<=zz[t][l][k]

#Constraint-17. Selecting proper ratcheting order for injection 
for t in T:
    for l in L:
        for k in L:
            for kr in L:
                if k>=l:
                    if kr>k:
                        prob+=zz[t][l][k]>=zz[t][l][kr]
                        
#Constraint-18. Selecting proper ratcheting order for withdrawal 
for t in T:
    for l in L:
        for k in L:
            if k<=l:
                prob+=mu[t][l][k]<=xx[t][l][k]
                
#Constraint-19. Selecting proper ratcheting order for withdrawal 
for t in T:
    for l in L:
        for k in L:
            for kr in L:
                if k<=l:
                    if kr<k:
                        prob+=xx[t][l][k]>=xx[t][l][kr]                        

#Constraint-20 & 21. Propertion of ratchet l has to be greater than 0
for t in T:
    for l in L:
         for k in L:
            prob+=lb[t][l][k]>=0
            prob+=mu[t][l][k]>=0

#Constraint-22. Inventory flow balance
for t in T:
    if t>0:
        b=t-1
        prob+= I[t]==I[b]+v[b]-w[b]

#Constraint-23 & 24. Inventory at the begining of first and last time period
prob+= I[Horizon-1]==0
prob+= I[0]==0

#################################################Solving with default cbc solver################################################
prob.solve()

#################################################printing the output############################################################
print("Max profit:", prob.objective.value())

########################################Writing the output in excel file#######################################################

#Extracting the required variables from the model output
for t in T: 
    vValue.append(str(v[t].varValue)) #Injection quantities
    wValue.append(str(w[t].varValue)) #withdrawal quantities
    if t==0:
        i=(v[t].varValue - w[t].varValue) #Inventory at the end of time period t
    else:
        i=i+(v[t].varValue - w[t].varValue) ##Inventory at the end of time period t
    iValue.append(i)
    g=(-(v[t].varValue)+(w[t].varValue))*p[t] #Monthly cashflow
    gValue.append(g)
    if (v[t].varValue + w[t].varValue)>0:
        Actions.append(v[t].varValue+w[t].varValue) #No. of storage actions (injection & withdrawal)
 
Index=[0 for i in range(Horizon)] 
for i in range(Horizon):
    Index[i]=ws1.cell(row=4, column=i+2).value #Indexing by time period
    
#Converting the variables output into pandas dataframes
df_vName= pd.DataFrame({'n':Index}).transpose()
df_vValue= pd.DataFrame({'Injection Amount(MMBtu):':vValue}).transpose()
df_wValue= pd.DataFrame({'Withdrawal Amount(MMBtu):':wValue}).transpose()
df_iValue= pd.DataFrame({'Inv. Level at end of period:':iValue}).transpose()
df_gValue= pd.DataFrame({'Value(-ve for purchase):':gValue}).transpose()
df_ContractValue= pd.DataFrame({'Approximate Contract Value:':[prob.objective.value()]}).transpose()
df_Actions= pd.DataFrame({'Number of Actions:': [len(Actions)]}).transpose()
df_AvgInv= pd.DataFrame({'Avg. Inv. Level (MMBtu):': [sum(iValue)/int(line1[1][10:])]}).transpose()

#Writing the dataframes in the excel file (Results.xlsx)
writer = pd.ExcelWriter('Results.xlsx', engine='xlsxwriter')
df_vName.to_excel(writer, sheet_name='Sheet1', startcol=1, startrow=0, header=False, index=False)
df_vValue.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=1, header=False, index=True)
df_wValue.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=2, header=False, index=True)
df_iValue.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=3, header=False, index=True)
df_gValue.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=4, header=False, index=True)
df_ContractValue.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=6, header=False, index=True)
df_Actions.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=7, header=False, index=True)
df_AvgInv.to_excel(writer, sheet_name='Sheet1', startcol=0, startrow=8, header=False, index=True)

#Formatting & saving the excel file
workbook= writer.book 
worksheet= writer.sheets['Sheet1']
border_format=workbook.add_format({'border':1, 'align':'left', 'font_size':10})
worksheet.conditional_format('A1:Y9',{'type' : 'no_blanks' , 'format' : border_format})                             
writer.save() 

##################################Printing the computational time#########################################
print("--- %s seconds ---" % (time.time() - start_time))    
######################################End of the code#####################################################

