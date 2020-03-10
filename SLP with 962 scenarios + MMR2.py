import time
start_time = time.time()
from pulp import *
from openpyxl import load_workbook

#For 962 scenarios#
wb=load_workbook(filename= 'Price Curves 12-5-18.xlsx')
ws=wb['Sheet2']

wb2=load_workbook(filename= 'X_Sol values for 962 clusters.xlsx') #For 962 Clusters
ws2=wb2['Sheet1']


M=10000000
Cap=1000000
cin=305000
cout=-457500
d=1
Horizon=24
N=962 #For 962 Clusters

T=list(range(Horizon))
S=list(range(N))

p=[[0 for t in range(Horizon)] for s in range(N)]
x_sol=[[0 for t in range(Horizon)] for s in range(N)]

for t in range(Horizon):
    for s in range(N):
        p[s][t]=ws.cell(row=t+2, column=s+2).value  #For 962 scenarios
for t in range(Horizon):
    for s in range(N):
        x_sol[s][t]=ws2.cell(row=t+1, column=s+1).value   #For 962 scenarios
prob= LpProblem("EL-NR", LpMinimize)
x= LpVariable.matrix("x", (T), None, None, LpContinuous)
I= LpVariable.matrix("I", (T), 0, None, LpContinuous)
r= LpVariable('r',  0, None, LpContinuous)
        
prob+= r
for s in S:
    prob+=lpSum([-p[s][t]*x[t] for t in T]) + r >= lpSum([-p[s][t]*x_sol[s][t] for t in T])

for t in T:
    prob+=cout<=x[t]
    prob+=x[t]<=cin
    prob+=I[t]<=Cap
    if t>0:
        prob+=I[t]==I[t-1]+x[t-1]

    prob+= I[Horizon-1]+x[Horizon-1]==I[0]
    prob+= I[0]==0  
    prob.solve()

print(prob.objective.value())

for v in prob.variables():
    if v.name[:1]=="x":
        print(v.name, "=", v.varValue)
        
print("--- %s seconds ---" % (time.time() - start_time))
